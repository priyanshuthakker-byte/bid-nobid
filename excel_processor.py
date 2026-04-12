"""
Excel Processor v3 - Reads T247 Excel and classifies tenders
FIXED: Added invalidate_rules_cache() so profile saves take effect immediately
FIXED: Added 'bid' key handling in classify_tender
"""

import re, json
from pathlib import Path
from typing import List, Dict
from datetime import datetime, date

PROFILE_PATH = Path(__file__).parent / "nascent_profile.json"

# Cache for bid rules — invalidated when profile is saved
_rules_cache = None

def invalidate_rules_cache():
    """Call this after saving nascent_profile.json to force fresh rule load."""
    global _rules_cache
    _rules_cache = None

def load_rules() -> Dict:
    global _rules_cache
    if _rules_cache is not None:
        return _rules_cache
    try:
        p = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
        _rules_cache = p.get("bid_rules", {})
        return _rules_cache
    except Exception:
        return {}

def classify_tender(brief: str, estimated_cost: float,
                    eligibility: str, checklist: str) -> Dict:
    rules = load_rules()

    dnb_keywords = [k.lower().strip() for k in rules.get("do_not_bid", [])]
    dnb_remarks = {k.lower(): v for k, v in rules.get("do_not_bid_remarks", {}).items()}
    cond_keywords = [k.lower().strip() for k in rules.get("conditional", [])]
    pref_keywords = [k.lower().strip() for k in rules.get("preferred_sectors", [])]
    bid_keywords = [k.lower().strip() for k in rules.get("bid", [])]  # FIXED: explicit BID list
    min_val = rules.get("min_project_value_cr", 0.25)
    max_val = rules.get("max_project_value_cr", 100)

    brief_lower = str(brief or "").lower().strip()
    full_text = " ".join([str(brief or ""), str(eligibility or ""), str(checklist or "")]).lower()

    # Rule 1: DO NOT BID
    for kw in dnb_keywords:
        if kw in brief_lower:
            if any(pk in brief_lower for pk in pref_keywords):
                continue
            remark = dnb_remarks.get(kw, "")
            for dk, dr in dnb_remarks.items():
                if dk in brief_lower:
                    remark = dr
                    break
            if not remark:
                remark = f"Matches NO-BID rule: '{kw}' — not Nascent's service domain."
            return {"verdict": "NO-BID", "verdict_color": "RED", "reason": remark}

    # Rule 2: Value too low
    if estimated_cost and 0 < estimated_cost < (min_val * 1_00_00_000):
        return {"verdict": "NO-BID", "verdict_color": "RED",
                "reason": f"Project value below Nascent minimum threshold of Rs. {min_val} Cr."}

    # Rule 3: Value too high
    if estimated_cost and estimated_cost > (max_val * 1_00_00_000):
        return {"verdict": "CONDITIONAL", "verdict_color": "AMBER",
                "reason": f"Project value exceeds Rs. {max_val} Cr. Verify turnover eligibility. Raise pre-bid query for MSME relaxation."}

    # Rule 4: CONDITIONAL triggers
    for kw in cond_keywords:
        if kw in full_text:
            return {"verdict": "CONDITIONAL", "verdict_color": "AMBER",
                    "reason": f"Conditional trigger: '{kw}' — verify eligibility and raise pre-bid query."}

    # Rule 5: Explicit BID keywords (FIXED — was dead because 'bid' key missing)
    if bid_keywords:
        matched_bid = [kw for kw in bid_keywords if kw in full_text]
        if matched_bid:
            return {"verdict": "BID", "verdict_color": "GREEN",
                    "reason": f"Explicit BID match: {', '.join(matched_bid[:3])}."}

    # Rule 6: Preferred sectors → BID
    matched = [kw for kw in pref_keywords if kw in full_text]
    if matched:
        return {"verdict": "BID", "verdict_color": "GREEN",
                "reason": f"Matches Nascent preferred sector: {', '.join(matched[:3])}."}

    # Rule 7: Corrigendum → REVIEW
    if "corrigendum" in brief_lower or "addendum" in brief_lower:
        return {"verdict": "REVIEW", "verdict_color": "BLUE",
                "reason": "Corrigendum/Addendum — check original tender first."}

    # Rule 8: Default
    return {"verdict": "REVIEW", "verdict_color": "BLUE",
            "reason": "Requires manual review — insufficient data to auto-classify."}

def days_left(deadline_str: str) -> int:
    if not deadline_str:
        return 999
    for fmt in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y"]:
        try:
            d = datetime.strptime(str(deadline_str).split()[0], fmt).date()
            return (d - date.today()).days
        except Exception:
            continue
    return 999

def deadline_status(dl: int) -> str:
    if dl < 0: return "EXPIRED"
    if dl == 0: return "TODAY"
    if dl <= 3: return "URGENT"
    if dl <= 7: return "THIS_WEEK"
    if dl <= 14: return "SOON"
    return "OK"

def process_excel(filepath: str) -> List[Dict]:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_tenders = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_gem = sheet_name.lower().strip() == "gem tenders" or sheet_name.lower().startswith("gem")

        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

        def col(*names):
            for name in names:
                for i, h in enumerate(headers):
                    if name in h:
                        return i
            return None

        idx = {
            "t247_id": col("t247 id", "t247id"),
            "ref_no": col("reference no", "ref no", "reference"),
            "brief": col("tender brief", "brief", "description", "title"),
            "cost": col("estimated cost", "cost", "value"),
            "deadline": col("deadline", "last date", "submission date"),
            "location": col("location", "state", "city"),
            "org": col("organization", "organisation", "dept", "department"),
            "doc_fee": col("document fee", "doc fee", "tender fee", "processing fee"),
            "emd": col("emd", "earnest"),
            "msme": col("msme exemption", "msme"),
            "startup": col("startup exemption", "startup"),
            "quantity": col("quantity"),
            "eligibility": col("eligibility criteria", "eligibility"),
            "checklist": col("checklist"),
            "bid_opening": col("bid opening date", "opening date"),
            "bid_validity": col("bid offer validity", "validity"),
            "ministry": col("ministry", "ministry/state"),
            "department": col("department name"),
            "office": col("office name"),
            "turnover_req": col("minimum average annual turnover", "turnover of the bidder"),
            "exp_years": col("years of past experience", "past experience required"),
            "contract_period": col("contract period"),
            "similar_cat": col("similar category"),
            "eval_method": col("evaluation method"),
            "pbg_pct": col("epbg percentage", "pbg percentage"),
            "mse_pref": col("mse purchase preference"),
            "type_of_bid": col("type of bid"),
        }

        for row in ws.iter_rows(min_row=2, values_only=True):
            t247_id = row[idx["t247_id"]] if idx["t247_id"] is not None else None
            if not t247_id:
                continue

            def cell(key):
                i = idx.get(key)
                return row[i] if i is not None and i < len(row) else ""

            def cs(key): return str(cell(key) or "").strip()

            brief = cs("brief")
            cost_raw = cell("cost")
            deadline = cs("deadline")
            org = cs("org")
            location = cs("location")
            emd = cs("emd")
            doc_fee = cs("doc_fee")
            msme = cs("msme")
            ref_no = cs("ref_no")
            elig = cs("eligibility")
            chklist = cs("checklist")

            try:
                cost = float(str(cost_raw).replace(",", "")) if cost_raw else 0
            except Exception:
                cost = 0

            classification = classify_tender(brief, cost, elig, chklist)
            dl = days_left(deadline)
            cost_cr = round(cost / 1_00_00_000, 2) if cost else 0

            tender = {
                "t247_id": str(t247_id),
                "ref_no": ref_no,
                "brief": brief[:200],
                "org_name": org,
                "location": location,
                "estimated_cost_raw": cost,
                "estimated_cost_cr": cost_cr,
                "deadline": deadline,
                "days_left": dl,
                "deadline_status": deadline_status(dl),
                "doc_fee": doc_fee,
                "emd": emd,
                "msme_exemption": msme,
                "startup_exemption": cs("startup"),
                "quantity": cs("quantity"),
                "eligibility": elig[:1000],
                "checklist": chklist[:1000],
                "is_gem": is_gem,
                "verdict": classification["verdict"],
                "verdict_color": classification["verdict_color"],
                "reason": classification["reason"],
                "bid_no_bid_done": False,
                "status": "Identified",
                "imported_at": datetime.now().isoformat(),
            }

            if is_gem:
                tender.update({
                    "bid_opening_date": cs("bid_opening"),
                    "bid_validity": cs("bid_validity"),
                    "ministry": cs("ministry"),
                    "department": cs("department"),
                    "office": cs("office"),
                    "turnover_required": cs("turnover_req"),
                    "experience_years": cs("exp_years"),
                    "contract_period": cs("contract_period"),
                    "similar_category": cs("similar_cat"),
                    "evaluation_method": cs("eval_method"),
                    "pbg_percentage": cs("pbg_pct"),
                    "mse_preference": cs("mse_pref"),
                    "type_of_bid": cs("type_of_bid"),
                })

            all_tenders.append(tender)

    return all_tenders

def quick_classify(brief: str, cost: float = 0, eligibility: str = "") -> Dict:
    return classify_tender(brief, cost, eligibility, "")
